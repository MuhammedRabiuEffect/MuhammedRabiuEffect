import qrcode
import os
import winsound
import openpyxl


xl_file = './CLASS REGISTER-1.xlsx'
wb_obj = openpyxl.load_workbook(xl_file)
sheet = wb_obj.active

qr = qrcode.QRCode(
    version=20,
    error_correction=qrcode.constants.ERROR_CORRECT_L,
    box_size=5,
    border=2
)

for row in range(2, 20):
    name = sheet.cell(column=1, row=row).value
    number = sheet.cell(column=2, row=row).value
    # qr.add_data(number)
    qr.add_data(str(number)[-9:])

    background_color = '#23dda0'
    img = qr.make_image(fill_color="red", back_color=background_color)
    # img.save(f'{name}.png')

    img.save(f'./QRS_V_XX/{name}.png')

    winsound.Beep(5000, 200)
    print(f'{name}.......{number} done')
